from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from django.db.models import Q, Count, Avg
from django.http import HttpResponse
from datetime import timedelta
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from .models import EODReport, ReportReview
from .forms import EODReportForm, ReportReviewForm, EODReportFilterForm
from .utils import get_week_date_range, is_weekend, get_week_display


@login_required
def dashboard_view(request):
    """Main dashboard - redirects based on user role"""
    if request.user.is_manager() or request.user.is_admin_user():
        return redirect('reports:manager_dashboard')
    else:
        return redirect('reports:employee_dashboard')


@login_required
def employee_dashboard_view(request):
    """Employee dashboard showing their reports"""
    # Get employee's reports
    reports = EODReport.objects.filter(employee=request.user).order_by('-report_date')[:10]

    # Check if today's report exists
    today = timezone.now().date()
    try:
        today_report = EODReport.objects.get(employee=request.user, report_date=today)
        today_report_exists = True
        can_edit_today_report = today_report.can_edit()
    except EODReport.DoesNotExist:
        today_report = None
        today_report_exists = False
        can_edit_today_report = False

    # Get current week date range
    week_start, week_end = get_week_date_range()

    # Get rejected reports that can be edited (within 7 days and under 3 attempts)
    rejected_editable_reports = []
    all_rejected = EODReport.objects.filter(employee=request.user, status='REJECTED')
    for report in all_rejected:
        if report.can_edit():
            rejected_editable_reports.append(report)

    # Get overall statistics
    total_reports = EODReport.objects.filter(employee=request.user).count()
    pending_reports = EODReport.objects.filter(employee=request.user, status='PENDING').count()
    approved_reports = EODReport.objects.filter(employee=request.user, status='APPROVED').count()

    # Get weekly statistics
    week_reports = EODReport.objects.filter(
        employee=request.user,
        report_date__gte=week_start,
        report_date__lte=week_end
    )
    week_total = week_reports.count()
    week_pending = week_reports.filter(status='PENDING').count()
    week_approved = week_reports.filter(status='APPROVED').count()
    week_rejected = week_reports.filter(status='REJECTED').count()

    context = {
        'reports': reports,
        'today_report': today_report,
        'today_report_exists': today_report_exists,
        'can_edit_today_report': can_edit_today_report,
        'today': today,
        'is_weekend': is_weekend(),
        'week_display': get_week_display(),
        'week_start': week_start,
        'week_end': week_end,
        # Rejected reports that can be resubmitted
        'rejected_editable_reports': rejected_editable_reports,
        # Overall stats
        'total_reports': total_reports,
        'pending_reports': pending_reports,
        'approved_reports': approved_reports,
        # Weekly stats
        'week_total': week_total,
        'week_pending': week_pending,
        'week_approved': week_approved,
        'week_rejected': week_rejected,
    }
    return render(request, 'reports/employee_dashboard.html', context)


@login_required
def submit_report_view(request, pk=None):
    """Submit or edit EOD report"""
    today = timezone.now().date()
    is_resubmission = False
    report = None
    is_edit = False

    # If pk is provided, we're editing a specific report (potentially a rejected one)
    if pk:
        try:
            report = EODReport.objects.get(pk=pk, employee=request.user)
            is_edit = True

            # Check if this is a resubmission of a rejected report
            if report.status == 'REJECTED':
                is_resubmission = True
                # Verify it can be edited (within 7 days and under 3 attempts)
                if not report.can_edit():
                    if report.resubmission_count >= 3:
                        messages.error(
                            request,
                            'Cannot edit this report. You have reached the maximum number of resubmissions (3 attempts).'
                        )
                    else:
                        messages.error(
                            request,
                            'Cannot edit this report. The 7-day edit window has expired.'
                        )
                    return redirect('reports:employee_dashboard')
            elif not report.can_edit():
                messages.error(
                    request,
                    f'Cannot edit this report. It has been {report.get_status_display().lower()} by your manager.'
                )
                return redirect('reports:employee_dashboard')
        except EODReport.DoesNotExist:
            messages.error(request, 'Report not found.')
            return redirect('reports:employee_dashboard')
    else:
        # Check if report for today already exists (only check weekday reports)
        try:
            # Try to find today's report only if today is a weekday
            if today.weekday() < 5:  # Monday-Friday
                report = EODReport.objects.get(employee=request.user, report_date=today)
                is_edit = True

                # Check if report can be edited (not approved/rejected)
                if not report.can_edit():
                    messages.error(
                        request,
                        f'Cannot edit this report. It has been {report.get_status_display().lower()} by your manager.'
                    )
                    return redirect('reports:employee_dashboard')
            else:
                # Weekend: Allow creating new reports for past weekdays
                report = None
                is_edit = False
        except EODReport.DoesNotExist:
            report = None
            is_edit = False

    if request.method == 'POST':
        form = EODReportForm(request.POST, instance=report, user=request.user)
        if form.is_valid():
            eod_report = form.save(commit=False)
            eod_report.employee = request.user

            # Handle resubmission logic
            if is_resubmission:
                eod_report.status = 'PENDING'
                eod_report.resubmission_count += 1
                eod_report.save()
                messages.success(
                    request,
                    f'Report resubmitted successfully! (Attempt {eod_report.resubmission_count}/3)'
                )
            else:
                eod_report.save()
                if is_edit:
                    messages.success(request, 'EOD report updated successfully!')
                else:
                    messages.success(request, 'EOD report submitted successfully!')
            return redirect('reports:employee_dashboard')
    else:
        form = EODReportForm(instance=report, user=request.user)

    # Get last review if exists (for displaying rejection comments)
    last_review = None
    if report and report.reviews.exists():
        last_review = report.reviews.latest('reviewed_at')

    context = {
        'form': form,
        'is_edit': is_edit,
        'report': report,
        'is_resubmission': is_resubmission,
        'last_review': last_review,
    }
    return render(request, 'reports/submit_report.html', context)


@login_required
def report_detail_view(request, pk):
    """View individual report details"""
    report = get_object_or_404(EODReport, pk=pk)

    # Check permissions
    if request.user.is_employee():
        if report.employee != request.user:
            messages.error(request, 'You do not have permission to view this report.')
            return redirect('reports:employee_dashboard')
    elif request.user.is_manager():
        # Manager can only view their team members' reports
        if report.employee.manager != request.user:
            messages.error(request, 'You do not have permission to view this report.')
            return redirect('reports:manager_dashboard')

    # Get all reviews for this report (ordered from oldest to newest)
    reviews = report.reviews.all().order_by('review_number')

    context = {
        'report': report,
        'reviews': reviews,
        'can_edit': report.can_edit() and report.employee == request.user,
        'remaining_attempts': report.remaining_resubmissions() if report.is_rejected() else 0,
    }
    return render(request, 'reports/report_detail.html', context)


@login_required
def manager_dashboard_view(request):
    """Manager dashboard for reviewing team reports"""
    if not (request.user.is_manager() or request.user.is_admin_user()):
        messages.error(request, 'You do not have permission to access this page.')
        return redirect('reports:employee_dashboard')

    # Get team members' reports
    if request.user.is_admin_user():
        reports = EODReport.objects.all()
        team_members = None
    else:
        team_members = request.user.get_team_members()
        reports = EODReport.objects.filter(employee__in=team_members)

    # Apply filters
    filter_form = EODReportFilterForm(request.GET)
    if filter_form.is_valid():
        date_from = filter_form.cleaned_data.get('date_from')
        date_to = filter_form.cleaned_data.get('date_to')
        status = filter_form.cleaned_data.get('status')
        employee = filter_form.cleaned_data.get('employee')

        if date_from:
            reports = reports.filter(report_date__gte=date_from)
        if date_to:
            reports = reports.filter(report_date__lte=date_to)
        if status:
            reports = reports.filter(status=status)
        if employee:
            reports = reports.filter(
                Q(employee__first_name__icontains=employee) |
                Q(employee__last_name__icontains=employee) |
                Q(employee__username__icontains=employee)
            )

    reports = reports.select_related('employee').order_by('-report_date', '-submitted_at')[:50]

    # Get current week date range
    week_start, week_end = get_week_date_range()

    # Get base queryset for statistics
    if request.user.is_admin_user():
        base_reports = EODReport.objects.all()
    else:
        base_reports = EODReport.objects.filter(employee__manager=request.user)

    # Get overall statistics
    pending_count = base_reports.filter(status='PENDING').count()
    total_count = base_reports.count()
    approved_count = base_reports.filter(status='APPROVED').count()

    # Get weekly statistics
    week_reports = base_reports.filter(
        report_date__gte=week_start,
        report_date__lte=week_end
    )
    week_total = week_reports.count()
    week_pending = week_reports.filter(status='PENDING').count()
    week_approved = week_reports.filter(status='APPROVED').count()
    week_rejected = week_reports.filter(status='REJECTED').count()

    context = {
        'reports': reports,
        'filter_form': filter_form,
        'is_weekend': is_weekend(),
        'week_display': get_week_display(),
        'week_start': week_start,
        'week_end': week_end,
        # Overall stats
        'pending_count': pending_count,
        'total_count': total_count,
        'approved_count': approved_count,
        # Weekly stats
        'week_total': week_total,
        'week_pending': week_pending,
        'week_approved': week_approved,
        'week_rejected': week_rejected,
        'team_members': team_members,
    }
    return render(request, 'reports/manager_dashboard.html', context)


@login_required
def review_report_view(request, pk):
    """Manager review and approve/reject report"""
    if not (request.user.is_manager() or request.user.is_admin_user()):
        messages.error(request, 'You do not have permission to review reports.')
        return redirect('reports:employee_dashboard')

    report = get_object_or_404(EODReport, pk=pk)

    # Check if manager has permission to review this report
    if request.user.is_manager() and report.employee.manager != request.user:
        messages.error(request, 'You can only review reports from your team members.')
        return redirect('reports:manager_dashboard')

    # Check if report is pending review
    if report.status != 'PENDING':
        messages.info(
            request,
            f'This report has already been {report.get_status_display().lower()}. '
            'It cannot be reviewed at this time.'
        )
        # Redirect to report detail instead of review page
        return redirect('reports:report_detail', pk=report.pk)

    # Get previous reviews to determine review number and check if this is a resubmission
    previous_reviews = report.reviews.all().order_by('review_number')
    is_resubmission = previous_reviews.exists()
    next_review_number = previous_reviews.count() + 1 if is_resubmission else 1

    # Always create a new review (never update existing ones to preserve history)
    review = None
    is_new_review = True

    if request.method == 'POST':
        # Double-check status hasn't changed since page load
        if report.status != 'PENDING':
            messages.error(request, 'This report has already been reviewed and cannot be modified.')
            return redirect('reports:manager_dashboard')

        form = ReportReviewForm(request.POST, instance=review, report=report)
        if form.is_valid():
            report_review = form.save(commit=False)
            report_review.report = report
            report_review.reviewer = request.user
            report_review.review_number = next_review_number
            report_review.save()

            # Update report status
            decision = form.cleaned_data.get('decision')
            report.status = decision
            report.save()

            action = 'approved' if decision == 'APPROVED' else 'rejected'
            review_iteration = f' (Review #{next_review_number})' if is_resubmission else ''
            messages.success(
                request,
                f'Report {action} successfully{review_iteration}! '
                'Employee can resubmit if rejected (within 7 days, max 3 attempts).'
            )
            return redirect('reports:manager_dashboard')
    else:
        form = ReportReviewForm(instance=review, report=report)

    context = {
        'form': form,
        'report': report,
        'is_new_review': is_new_review,
        'is_resubmission': is_resubmission,
        'previous_reviews': previous_reviews,
        'next_review_number': next_review_number,
    }
    return render(request, 'reports/review_report.html', context)


@login_required
def my_reports_view(request):
    """View all reports for the current user"""
    reports = EODReport.objects.filter(employee=request.user).order_by('-report_date')

    # Apply date filters
    filter_form = EODReportFilterForm(request.GET)
    if filter_form.is_valid():
        date_from = filter_form.cleaned_data.get('date_from')
        date_to = filter_form.cleaned_data.get('date_to')
        status = filter_form.cleaned_data.get('status')

        if date_from:
            reports = reports.filter(report_date__gte=date_from)
        if date_to:
            reports = reports.filter(report_date__lte=date_to)
        if status:
            reports = reports.filter(status=status)

    context = {
        'reports': reports,
        'filter_form': filter_form,
    }
    return render(request, 'reports/my_reports.html', context)


@login_required
def export_reports_excel(request):
    """Export filtered reports to Excel for managers"""
    if not (request.user.is_manager() or request.user.is_admin_user()):
        messages.error(request, 'You do not have permission to export reports.')
        return redirect('reports:employee_dashboard')

    # Get team members' reports (same logic as manager dashboard)
    if request.user.is_admin_user():
        reports = EODReport.objects.all()
    else:
        team_members = request.user.get_team_members()
        reports = EODReport.objects.filter(employee__in=team_members)

    # Apply filters (same as manager dashboard)
    filter_form = EODReportFilterForm(request.GET)
    if filter_form.is_valid():
        date_from = filter_form.cleaned_data.get('date_from')
        date_to = filter_form.cleaned_data.get('date_to')
        status = filter_form.cleaned_data.get('status')
        employee = filter_form.cleaned_data.get('employee')

        if date_from:
            reports = reports.filter(report_date__gte=date_from)
        if date_to:
            reports = reports.filter(report_date__lte=date_to)
        if status:
            reports = reports.filter(status=status)
        if employee:
            reports = reports.filter(
                Q(employee__first_name__icontains=employee) |
                Q(employee__last_name__icontains=employee) |
                Q(employee__username__icontains=employee)
            )

    # Remove the limit for export (get all matching reports, not just 50)
    reports = reports.select_related('employee').order_by('-report_date', '-submitted_at')

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "EOD Reports"

    # Define headers
    headers = [
        'Employee Name',
        'Department',
        'Report Date',
        'Project Name',
        'Tasks Completed',
        'Hours Worked',
        'Blockers/Issues',
        'Next Day Plan',
        'Status',
        'Resubmission Count',
        'Submitted At',
        'Last Review Comments'
    ]

    # Style for header row
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Helper function to strip HTML tags
    def strip_html(text):
        """Remove HTML tags from text"""
        if not text:
            return ''
        clean = re.compile('<.*?>')
        return re.sub(clean, '', str(text)).strip()

    # Style for status cells
    status_fills = {
        'APPROVED': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        'PENDING': PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        'REJECTED': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    }

    # Write data rows
    for row_num, report in enumerate(reports, 2):
        # Get last review comments if any
        last_review = report.reviews.order_by('-reviewed_at').first()
        review_comments = last_review.comments if last_review else 'No review yet'

        # Prepare row data
        row_data = [
            report.employee.get_full_name() or report.employee.username,
            report.employee.department or 'N/A',
            report.report_date.strftime('%Y-%m-%d'),
            report.project_name,
            strip_html(report.tasks_completed),
            float(report.hours_worked),
            strip_html(report.blockers_issues) if report.blockers_issues else 'None',
            strip_html(report.next_day_plan),
            report.get_status_display(),
            report.resubmission_count,
            report.submitted_at.strftime('%Y-%m-%d %H:%M'),
            strip_html(review_comments)
        ]

        # Write row
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = Alignment(vertical="top", wrap_text=True)

            # Apply status color
            if col_num == 9:  # Status column
                cell.fill = status_fills.get(report.status, PatternFill())

    # Auto-size columns
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        for cell in ws[column_letter]:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        # Set width (with limits)
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = max(adjusted_width, 12)

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    # Generate filename with current date
    filename = f'EOD_Reports_{timezone.now().strftime("%Y-%m-%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Save workbook to response
    wb.save(response)

    return response
